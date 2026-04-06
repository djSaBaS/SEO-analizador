# Importa utilidades de fechas para cálculos de comparación temporal.
from datetime import date, datetime, timedelta

# Importa Path para gestionar salidas de archivos de forma segura.
from pathlib import Path

# Importa tipado para estructuras de datos del informe.
from typing import Any

# Importa Workbook para exportar datasets GA4 a Excel.
from openpyxl import Workbook

# Importa estilos para mejorar la legibilidad del Excel premium.
from openpyxl.styles import Font, PatternFill

# Importa utilidades de maquetación PDF.
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

# Importa configuración global del proyecto.
from seo_auditor.config import Configuracion

# Define dimensiones y métricas base para consultas agregadas del informe.
METRICAS_KPI = [
    "totalUsers",
    "newUsers",
    "activeUsers",
    "sessions",
    "eventCount",
    "conversions",
    "userEngagementDuration",
]

# Define umbrales de insights para evitar números mágicos.
INSIGHT_MIN_SESIONES_SIN_CONVERSION = 30.0
INSIGHT_MAX_CONVERSIONES_SIN_CONVERSION = 0.0
INSIGHT_MIN_SESIONES_REBOTE = 30.0
INSIGHT_MIN_REBOTE = 0.7
INSIGHT_MIN_SESIONES_ALTO_VALOR = 40.0
INSIGHT_MIN_CONVERSIONES_ALTO_VALOR = 2.0
INSIGHT_MUESTRA_URLS = 3

# Define límites de exportación para hoja Excel GA4.
EXCEL_MAX_FILAS_PAISES = 50
EXCEL_MAX_FILAS_COMUNIDADES = 50
EXCEL_MAX_FILAS_CIUDADES = 50
EXCEL_MAX_FILAS_ADQUISICION = 50
EXCEL_MAX_FILAS_REFERIDOS = 50
EXCEL_MAX_FILAS_LANDINGS = 80

# Define dimensiones y límites de render para gráficos PDF.
PDF_GRAFICO_ANCHO = 500
PDF_GRAFICO_ALTO = 280
PNG_GRAFICO_ANCHO = 1200
PNG_GRAFICO_ALTO = 700


# Calcula fecha equivalente restando un año con compatibilidad bisiesta.
def _restar_un_anio(fecha: date) -> date:
    """Resta un año ajustando 29 de febrero al 28 de febrero cuando aplique."""

    # Intenta reemplazar directamente el año para fechas válidas.
    try:
        # Devuelve fecha con año restado en caso normal.
        return fecha.replace(year=fecha.year - 1)

    # Corrige únicamente el caso bisiesto de 29 de febrero.
    except ValueError:
        # Devuelve 28 de febrero del año anterior cuando no existe el 29.
        return fecha.replace(month=2, day=28, year=fecha.year - 1)


# Calcula el rango de comparación para periodo anterior o año anterior.
def _resolver_comparacion(fecha_desde: str, fecha_hasta: str, modo: str) -> tuple[str, str]:
    """Devuelve fechas de comparación para el modo solicitado."""

    # Convierte la fecha inicial a objeto date.
    inicio = datetime.strptime(fecha_desde, "%Y-%m-%d").date()

    # Convierte la fecha final a objeto date.
    fin = datetime.strptime(fecha_hasta, "%Y-%m-%d").date()

    # Resuelve comparación frente al mismo periodo del año anterior.
    if modo == "anio-anterior":
        # Devuelve rango desplazado exactamente un año con control de bisiestos.
        return _restar_un_anio(inicio).isoformat(), _restar_un_anio(fin).isoformat()

    # Calcula longitud del periodo base de forma inclusiva.
    dias = (fin - inicio).days + 1

    # Aplica comparación frente al periodo inmediatamente anterior.
    return (inicio - timedelta(days=dias)).isoformat(), (inicio - timedelta(days=1)).isoformat()


# Crea un cliente autenticado de Analytics Data API.
def _crear_cliente(ruta_credenciales: str) -> Any:
    """Inicializa cliente oficial de Google Analytics Data API."""

    # Importa cliente oficial dentro de la función para degradación elegante.
    from google.analytics.data_v1beta import BetaAnalyticsDataClient

    # Retorna cliente autenticado usando service account local.
    return BetaAnalyticsDataClient.from_service_account_file(ruta_credenciales)


# Ejecuta una consulta GA4 genérica devolviendo filas planas.
def _consultar_ga4(
    cliente: Any,
    property_id: str,
    fecha_desde: str,
    fecha_hasta: str,
    dimensiones: list[str],
    metricas: list[str],
    limite: int = 1000,
) -> list[dict[str, Any]]:
    """Consulta GA4 y mapea filas a diccionarios serializables."""

    # Importa tipos de request de la librería oficial de GA4.
    from google.analytics.data_v1beta.types import DateRange, Dimension, Metric, RunReportRequest

    # Construye solicitud parametrizada con dimensiones y métricas.
    request = RunReportRequest(
        property=f"properties/{property_id}",
        dimensions=[Dimension(name=nombre) for nombre in dimensiones],
        metrics=[Metric(name=nombre) for nombre in metricas],
        date_ranges=[DateRange(start_date=fecha_desde, end_date=fecha_hasta)],
        limit=limite,
    )

    # Ejecuta consulta remota contra GA4.
    respuesta = cliente.run_report(request)

    # Inicializa colección de filas normalizadas.
    filas: list[dict[str, Any]] = []

    # Recorre cada fila de la respuesta con fallback seguro.
    for fila in list(respuesta.rows or []):
        # Inicializa diccionario fila para salida.
        item: dict[str, Any] = {}

        # Copia cada dimensión por índice.
        for indice, nombre_dimension in enumerate(dimensiones):
            # Guarda valor de dimensión como texto plano.
            item[nombre_dimension] = fila.dimension_values[indice].value if fila.dimension_values else ""

        # Copia cada métrica por índice con parseo a float.
        for indice, nombre_metrica in enumerate(metricas):
            # Convierte valor de métrica a flotante para cálculos posteriores.
            item[nombre_metrica] = float(fila.metric_values[indice].value or 0.0) if fila.metric_values else 0.0

        # Añade fila normalizada a la colección final.
        filas.append(item)

    # Devuelve filas normalizadas para reporting.
    return filas


# Convierte duración en segundos a texto legible.
def _formatear_segundos(segundos: float) -> str:
    """Formatea segundos como HH:MM:SS para lectura ejecutiva."""

    # Redondea segundos para evitar ruido decimal visual.
    total = int(round(segundos))

    # Calcula horas, minutos y segundos residuales.
    horas, resto = divmod(total, 3600)
    minutos, segundos_restantes = divmod(resto, 60)

    # Devuelve tiempo formateado con ceros iniciales.
    return f"{horas:02d}:{minutos:02d}:{segundos_restantes:02d}"


# Construye insights automáticos sobre landings de alto riesgo/valor.
def _construir_insights(landings: list[dict[str, Any]]) -> list[str]:
    """Detecta patrones de negocio accionables en páginas de entrada."""

    # Inicializa lista de insights para el informe.
    insights: list[str] = []

    # Detecta páginas con tráfico y cero conversiones.
    sin_conversion = [
        fila
        for fila in landings
        if fila.get("sessions", 0.0) >= INSIGHT_MIN_SESIONES_SIN_CONVERSION
        and fila.get("conversions", 0.0) <= INSIGHT_MAX_CONVERSIONES_SIN_CONVERSION
    ]

    # Añade insight si existe al menos una página candidata.
    if sin_conversion:
        # Construye muestra compacta de URLs con mayor impacto potencial.
        muestra = ", ".join(
            str(fila.get("landingPagePlusQueryString", "")) for fila in sin_conversion[:INSIGHT_MUESTRA_URLS]
        )

        # Registra insight de oportunidad de conversión.
        insights.append(f"Páginas con tráfico pero sin conversión: {muestra}.")

    # Detecta páginas con rebote alto en muestra suficiente.
    alto_rebote = [
        fila
        for fila in landings
        if fila.get("sessions", 0.0) >= INSIGHT_MIN_SESIONES_REBOTE
        and fila.get("bounceRate", 0.0) >= INSIGHT_MIN_REBOTE
    ]

    # Añade insight si existen páginas con rebote alto.
    if alto_rebote:
        # Construye muestra compacta de páginas con peor rebote.
        muestra = ", ".join(
            str(fila.get("landingPagePlusQueryString", "")) for fila in alto_rebote[:INSIGHT_MUESTRA_URLS]
        )

        # Registra insight de fricción UX/contenido.
        insights.append(f"Páginas con alto rebote detectadas: {muestra}.")

    # Detecta páginas con valor alto por combinación de sesiones y conversiones.
    alto_valor = [
        fila
        for fila in landings
        if fila.get("sessions", 0.0) >= INSIGHT_MIN_SESIONES_ALTO_VALOR
        and fila.get("conversions", 0.0) >= INSIGHT_MIN_CONVERSIONES_ALTO_VALOR
    ]

    # Añade insight de valor si hay páginas destacadas.
    if alto_valor:
        # Construye muestra de landings de alto valor.
        muestra = ", ".join(
            str(fila.get("landingPagePlusQueryString", "")) for fila in alto_valor[:INSIGHT_MUESTRA_URLS]
        )

        # Registra insight de replicación estratégica.
        insights.append(f"Páginas con alto valor para escalar: {muestra}.")

    # Devuelve insights finales o mensaje neutro cuando no haya señal.
    return insights or ["No se detectaron patrones críticos automáticos en el rango analizado."]


# Consulta todos los datasets necesarios para el informe premium.
def _cargar_datasets_ga4(
    cliente: Any, property_id: str, fecha_desde: str, fecha_hasta: str, fecha_comp_desde: str, fecha_comp_hasta: str
) -> dict[str, list[dict[str, Any]]]:
    """Agrupa consultas GA4 para simplificar la orquestación del informe."""

    # Devuelve todos los datasets en un único diccionario.
    return {
        "kpi_actual": _consultar_ga4(cliente, property_id, fecha_desde, fecha_hasta, [], METRICAS_KPI, 1),
        "kpi_comparado": _consultar_ga4(cliente, property_id, fecha_comp_desde, fecha_comp_hasta, [], METRICAS_KPI, 1),
        "paises": _consultar_ga4(cliente, property_id, fecha_desde, fecha_hasta, ["country"], ["totalUsers"], 250),
        "comunidades": _consultar_ga4(cliente, property_id, fecha_desde, fecha_hasta, ["region"], ["totalUsers"], 250),
        "ciudades": _consultar_ga4(cliente, property_id, fecha_desde, fecha_hasta, ["city"], ["totalUsers"], 250),
        "dispositivos": _consultar_ga4(
            cliente, property_id, fecha_desde, fecha_hasta, ["deviceCategory"], ["sessions"], 10
        ),
        "navegadores": _consultar_ga4(cliente, property_id, fecha_desde, fecha_hasta, ["browser"], ["sessions"], 20),
        "adquisicion_actual": _consultar_ga4(
            cliente, property_id, fecha_desde, fecha_hasta, ["sessionDefaultChannelGroup"], ["sessions"], 30
        ),
        "adquisicion_comparada": _consultar_ga4(
            cliente, property_id, fecha_comp_desde, fecha_comp_hasta, ["sessionDefaultChannelGroup"], ["sessions"], 30
        ),
        "referidos": _consultar_ga4(
            cliente, property_id, fecha_desde, fecha_hasta, ["sessionSourceMedium"], ["sessions", "conversions"], 50
        ),
        "social": _consultar_ga4(
            cliente, property_id, fecha_desde, fecha_hasta, ["sessionDefaultChannelGroup"], ["sessions"], 30
        ),
        "landings": _consultar_ga4(
            cliente,
            property_id,
            fecha_desde,
            fecha_hasta,
            ["landingPagePlusQueryString"],
            ["sessions", "totalUsers", "bounceRate", "conversions"],
            150,
        ),
    }


# Calcula kpis y variaciones para reutilizar en todas las exportaciones.
def _calcular_kpis(
    kpi_actual: list[dict[str, Any]], kpi_comparado: list[dict[str, Any]]
) -> tuple[dict[str, float], dict[str, float]]:
    """Calcula KPIs requeridos y variaciones frente al periodo comparado."""

    # Obtiene fila única de KPI actual con fallback vacío.
    kpi_a = kpi_actual[0] if kpi_actual else {nombre: 0.0 for nombre in METRICAS_KPI}

    # Obtiene fila única de KPI comparado con fallback vacío.
    kpi_c = kpi_comparado[0] if kpi_comparado else {nombre: 0.0 for nombre in METRICAS_KPI}

    # Deriva usuarios recurrentes como activos menos nuevos.
    usuarios_recurrentes = max(float(kpi_a.get("activeUsers", 0.0)) - float(kpi_a.get("newUsers", 0.0)), 0.0)

    # Deriva usuarios recurrentes comparativos para delta.
    usuarios_recurrentes_comp = max(float(kpi_c.get("activeUsers", 0.0)) - float(kpi_c.get("newUsers", 0.0)), 0.0)

    # Construye mapa de KPIs requeridos por negocio.
    kpis = {
        "usuarios": float(kpi_a.get("totalUsers", 0.0)),
        "usuarios_nuevos": float(kpi_a.get("newUsers", 0.0)),
        "usuarios_recurrentes": usuarios_recurrentes,
        "sesiones": float(kpi_a.get("sessions", 0.0)),
        "eventos": float(kpi_a.get("eventCount", 0.0)),
        "conversiones": float(kpi_a.get("conversions", 0.0)),
        "engagement_time": float(kpi_a.get("userEngagementDuration", 0.0)),
    }

    # Construye mapa de KPIs comparativos.
    kpis_comp = {
        "usuarios": float(kpi_c.get("totalUsers", 0.0)),
        "usuarios_nuevos": float(kpi_c.get("newUsers", 0.0)),
        "usuarios_recurrentes": usuarios_recurrentes_comp,
        "sesiones": float(kpi_c.get("sessions", 0.0)),
        "eventos": float(kpi_c.get("eventCount", 0.0)),
        "conversiones": float(kpi_c.get("conversions", 0.0)),
        "engagement_time": float(kpi_c.get("userEngagementDuration", 0.0)),
    }

    # Calcula variaciones porcentuales con protección ante división por cero.
    variaciones = {
        nombre: ((valor - kpis_comp[nombre]) / kpis_comp[nombre] * 100.0) if kpis_comp[nombre] > 0 else 0.0
        for nombre, valor in kpis.items()
    }

    # Devuelve KPIs y variaciones para reutilización en exportaciones.
    return kpis, variaciones


# Genera figuras plotly principales para el informe premium.
def _generar_graficos(datasets: dict[str, list[dict[str, Any]]]) -> tuple[dict[str, Any], dict[str, Any]]:
    """Construye gráficos y dataframes base a partir de datasets GA4."""

    # Importa librerías de visualización de manera perezosa.
    import pandas as pd
    import plotly.express as px
    import plotly.graph_objects as go

    # Convierte datasets a DataFrame para visualizaciones.
    dataframes = {
        "paises": pd.DataFrame(datasets.get("paises", [])),
        "comunidades": pd.DataFrame(datasets.get("comunidades", [])),
        "ciudades": pd.DataFrame(datasets.get("ciudades", [])),
        "dispositivos": pd.DataFrame(datasets.get("dispositivos", [])),
        "navegadores": pd.DataFrame(datasets.get("navegadores", [])),
        "adq_actual": pd.DataFrame(datasets.get("adquisicion_actual", [])),
        "adq_comp": pd.DataFrame(datasets.get("adquisicion_comparada", [])),
        "referidos": pd.DataFrame(datasets.get("referidos", [])),
        "landings": pd.DataFrame(datasets.get("landings", [])),
    }

    # Inicializa mapa de gráficos para salida.
    graficos: dict[str, Any] = {}

    # Crea mapa mundial por país cuando haya datos disponibles.
    if not dataframes["paises"].empty:
        graficos["mapa_mundial"] = px.choropleth(
            dataframes["paises"],
            locations="country",
            locationmode="country names",
            color="totalUsers",
            title="Usuarios por país",
            color_continuous_scale="Blues",
        )

    # Crea mapa de comunidades con enfoque regional de GA4.
    if not dataframes["comunidades"].empty:
        graficos["mapa_comunidades"] = px.bar(
            dataframes["comunidades"].sort_values("totalUsers", ascending=False).head(20),
            x="region",
            y="totalUsers",
            title="Usuarios por comunidad autónoma (región GA4)",
        )

    # Crea gráfico de barras para top ciudades.
    if not dataframes["ciudades"].empty:
        top_ciudades = dataframes["ciudades"].sort_values("totalUsers", ascending=False).head(15)
        graficos["top_ciudades"] = px.bar(top_ciudades, x="city", y="totalUsers", title="Top ciudades por usuarios")

    # Crea donut para distribución de dispositivos.
    if not dataframes["dispositivos"].empty:
        graficos["dispositivos"] = px.pie(
            dataframes["dispositivos"],
            names="deviceCategory",
            values="sessions",
            hole=0.5,
            title="Distribución por dispositivo",
        )

    # Crea barras horizontales para navegadores.
    if not dataframes["navegadores"].empty:
        top_navegadores = dataframes["navegadores"].sort_values("sessions", ascending=True).tail(12)
        graficos["navegadores"] = px.bar(
            top_navegadores, x="sessions", y="browser", orientation="h", title="Navegadores por sesiones"
        )

    # Crea barras comparativas para adquisición por canal.
    if not dataframes["adq_actual"].empty or not dataframes["adq_comp"].empty:
        comparativo = (
            dataframes["adq_actual"]
            .merge(
                dataframes["adq_comp"],
                how="outer",
                on="sessionDefaultChannelGroup",
                suffixes=("_actual", "_comparado"),
            )
            .fillna(0.0)
        )

        # Ordena por sesiones del periodo actual para legibilidad.
        comparativo = comparativo.sort_values("sessions_actual", ascending=False)

        # Construye figura comparativa agrupada por canal.
        graficos["adquisicion"] = go.Figure(
            data=[
                go.Bar(
                    name="Periodo actual", x=comparativo["sessionDefaultChannelGroup"], y=comparativo["sessions_actual"]
                ),
                go.Bar(
                    name="Periodo comparado",
                    x=comparativo["sessionDefaultChannelGroup"],
                    y=comparativo["sessions_comparado"],
                ),
            ]
        )
        graficos["adquisicion"].update_layout(barmode="group", title="Adquisición por canal")

    # Devuelve gráficos y dataframes para reutilización.
    return graficos, dataframes


# Exporta informe HTML interactivo con secciones ejecutivas.
def _exportar_html(
    ruta_html: Path,
    cliente_nombre: str,
    gestor: str,
    fecha_desde: str,
    fecha_hasta: str,
    fecha_comp_desde: str,
    fecha_comp_hasta: str,
    kpis: dict[str, float],
    variaciones: dict[str, float],
    insights: list[str],
    graficos: dict[str, Any],
    dataframes: dict[str, Any],
    social: list[dict[str, Any]],
    provincia: str,
) -> None:
    """Escribe versión HTML interactiva del informe premium."""

    # Filtra social orgánico y paid desde canal por defecto.
    social_organico = [fila for fila in social if "Organic Social" in str(fila.get("sessionDefaultChannelGroup", ""))]
    social_paid = [fila for fila in social if "Paid Social" in str(fila.get("sessionDefaultChannelGroup", ""))]

    # Construye contenido HTML interactivo.
    bloques_html = [
        f"<h1>Informe GA4 Premium · {cliente_nombre}</h1>",
        (
            f"<p>Gestor: {gestor} | Periodo: {fecha_desde} a {fecha_hasta} "
            f"| Comparación: {fecha_comp_desde} a {fecha_comp_hasta}</p>"
        ),
        "<h2>KPIs generales</h2>",
        "<ul>",
    ]

    # Añade KPIs con variación e indicador visual.
    for nombre, valor in kpis.items():
        # Calcula valor de variación del KPI actual.
        variacion = variaciones.get(nombre, 0.0)

        # Calcula flecha de tendencia según variación.
        flecha = "↑" if variacion >= 0 else "↓"

        # Formatea valor para engagement time en formato horario.
        valor_mostrado = _formatear_segundos(valor) if nombre == "engagement_time" else f"{valor:,.0f}"

        # Añade elemento de lista al bloque HTML.
        bloques_html.append(f"<li><b>{nombre}</b>: {valor_mostrado} ({flecha} {variacion:.2f}%)</li>")

    # Cierra lista de KPIs en HTML.
    bloques_html.append("</ul>")

    # Añade insights automáticos al HTML.
    bloques_html.append("<h2>Insights automáticos</h2><ul>")
    for insight in insights:
        bloques_html.append(f"<li>{insight}</li>")
    bloques_html.append("</ul>")

    # Inserta cada gráfico en modo interactivo cuando exista.
    for titulo, figura in graficos.items():
        bloques_html.append(f"<h2>{titulo.replace('_', ' ').title()}</h2>")
        bloques_html.append(figura.to_html(full_html=False, include_plotlyjs="cdn"))

    # Inserta tabla de referidos cuando haya filas disponibles.
    if not dataframes["referidos"].empty:
        bloques_html.append("<h2>Referidos (source / medium)</h2>")
        bloques_html.append(
            dataframes["referidos"].sort_values("sessions", ascending=False).head(20).to_html(index=False)
        )

    # Inserta tabla de landings cuando haya filas disponibles.
    if not dataframes["landings"].empty:
        bloques_html.append("<h2>Landing pages</h2>")
        bloques_html.append(
            dataframes["landings"].sort_values("sessions", ascending=False).head(30).to_html(index=False)
        )

    # Inserta resumen de social orgánico y paid.
    bloques_html.append("<h2>Redes sociales</h2>")
    bloques_html.append(f"<p>Orgánico: {sum(float(f.get('sessions', 0.0)) for f in social_organico):,.0f} sesiones</p>")
    bloques_html.append(f"<p>Paid: {sum(float(f.get('sessions', 0.0)) for f in social_paid):,.0f} sesiones</p>")

    # Añade bloque opcional de provincia para trazar foco local.
    if provincia.strip() and not dataframes["ciudades"].empty:
        # Filtra filas de ciudad que contengan el texto de provincia.
        filtro = dataframes["ciudades"][dataframes["ciudades"]["city"].str.contains(provincia, case=False, na=False)]

        # Añade resumen provincial al HTML.
        bloques_html.append(f"<h2>Detalle provincial: {provincia}</h2>")
        bloques_html.append(
            filtro.to_html(index=False) if not filtro.empty else "<p>No hay ciudades para la provincia indicada.</p>"
        )

    # Escribe el informe HTML completo en disco.
    ruta_html.write_text("\n".join(bloques_html), encoding="utf-8")


# Exporta informe Excel con Dashboard y hoja GA4 de detalle.
def _exportar_excel(
    ruta_excel: Path,
    cliente_nombre: str,
    fecha_desde: str,
    fecha_hasta: str,
    kpis: dict[str, float],
    variaciones: dict[str, float],
    datasets: dict[str, list[dict[str, Any]]],
) -> None:
    """Escribe versión Excel del informe premium con hoja GA4 específica."""

    # Construye libro Excel con hoja específica de GA4.
    libro = Workbook()
    hoja_dashboard = libro.active
    hoja_dashboard.title = "Dashboard"
    hoja_ga4 = libro.create_sheet("GA4")

    # Escribe cabecera visual del dashboard.
    hoja_dashboard["A1"] = "Dashboard GA4 Premium"
    hoja_dashboard["A1"].font = Font(size=16, bold=True)
    hoja_dashboard["A2"] = f"Cliente: {cliente_nombre}"
    hoja_dashboard["A3"] = f"Periodo: {fecha_desde} a {fecha_hasta}"

    # Vuelca KPIs en dashboard principal requerido.
    fila_dashboard = 5
    for nombre, valor in kpis.items():
        hoja_dashboard.cell(row=fila_dashboard, column=1, value=nombre)
        hoja_dashboard.cell(row=fila_dashboard, column=2, value=float(valor))
        hoja_dashboard.cell(row=fila_dashboard, column=3, value=float(variaciones.get(nombre, 0.0)))
        fila_dashboard += 1

    # Define cabeceras de hoja GA4 de detalle.
    hoja_ga4.append(["seccion", "dimension", "metrica_1", "metrica_2", "metrica_3", "metrica_4"])

    # Añade filas de países a la hoja GA4.
    for fila in datasets.get("paises", [])[:EXCEL_MAX_FILAS_PAISES]:
        hoja_ga4.append(["audiencia_pais", fila.get("country", ""), fila.get("totalUsers", 0.0), "", "", ""])

    # Añade filas de comunidad a la hoja GA4.
    for fila in datasets.get("comunidades", [])[:EXCEL_MAX_FILAS_COMUNIDADES]:
        hoja_ga4.append(["audiencia_comunidad", fila.get("region", ""), fila.get("totalUsers", 0.0), "", "", ""])

    # Añade filas de ciudad a la hoja GA4.
    for fila in datasets.get("ciudades", [])[:EXCEL_MAX_FILAS_CIUDADES]:
        hoja_ga4.append(["audiencia_ciudad", fila.get("city", ""), fila.get("totalUsers", 0.0), "", "", ""])

    # Añade filas de adquisición a la hoja GA4.
    for fila in datasets.get("adquisicion_actual", [])[:EXCEL_MAX_FILAS_ADQUISICION]:
        hoja_ga4.append(
            ["adquisicion", fila.get("sessionDefaultChannelGroup", ""), fila.get("sessions", 0.0), "", "", ""]
        )

    # Añade filas de referidos a la hoja GA4.
    for fila in datasets.get("referidos", [])[:EXCEL_MAX_FILAS_REFERIDOS]:
        hoja_ga4.append(
            [
                "referidos",
                fila.get("sessionSourceMedium", ""),
                fila.get("sessions", 0.0),
                fila.get("conversions", 0.0),
                "",
                "",
            ]
        )

    # Añade filas de landing pages a la hoja GA4.
    for fila in datasets.get("landings", [])[:EXCEL_MAX_FILAS_LANDINGS]:
        hoja_ga4.append(
            [
                "landing_pages",
                fila.get("landingPagePlusQueryString", ""),
                fila.get("sessions", 0.0),
                fila.get("totalUsers", 0.0),
                fila.get("bounceRate", 0.0),
                fila.get("conversions", 0.0),
            ]
        )

    # Estiliza fila de cabeceras en hoja de detalle.
    for celda in hoja_ga4[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(fill_type="solid", fgColor="1F4E78")

    # Guarda archivo Excel del informe premium.
    libro.save(ruta_excel)


# Exporta informe PDF estático e intenta incluir gráficos como imágenes.
def _exportar_pdf(
    ruta_pdf: Path,
    carpeta_salida: Path,
    cliente_nombre: str,
    gestor: str,
    fecha_desde: str,
    fecha_hasta: str,
    kpis: dict[str, float],
    variaciones: dict[str, float],
    insights: list[str],
    graficos: dict[str, Any],
) -> None:
    """Escribe versión PDF del informe premium con degradación por gráfico."""

    # Inicializa estructura del PDF premium.
    documento = SimpleDocTemplate(str(ruta_pdf), pagesize=A4)
    estilos = getSampleStyleSheet()
    contenido = []

    # Añade portada básica del PDF.
    contenido.append(Paragraph("Informe GA4 Premium", estilos["Title"]))
    contenido.append(Paragraph(f"Cliente: {cliente_nombre}", estilos["Normal"]))
    contenido.append(Paragraph(f"Gestor: {gestor}", estilos["Normal"]))
    contenido.append(Paragraph(f"Periodo: {fecha_desde} a {fecha_hasta}", estilos["Normal"]))
    contenido.append(Spacer(1, 12))

    # Construye tabla de KPIs para PDF.
    tabla_kpis = [["KPI", "Valor", "% variación"]]
    for nombre, valor in kpis.items():
        tabla_kpis.append(
            [
                nombre,
                _formatear_segundos(valor) if nombre == "engagement_time" else f"{valor:,.0f}",
                f"{variaciones.get(nombre, 0.0):.2f}%",
            ]
        )

    # Añade tabla de KPIs al PDF.
    tabla = Table(tabla_kpis)
    tabla.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ]
        )
    )
    contenido.append(tabla)
    contenido.append(Spacer(1, 12))

    # Añade insights automáticos al PDF.
    contenido.append(Paragraph("Insights automáticos", estilos["Heading2"]))
    for insight in insights:
        contenido.append(Paragraph(f"• {insight}", estilos["Normal"]))

    # Intenta renderizar gráficos estáticos a PNG para el PDF.
    for nombre, figura in graficos.items():
        try:
            # Define ruta de salida por gráfico para imagen estática.
            ruta_imagen = carpeta_salida / f"{nombre}.png"

            # Renderiza imagen PNG desde figura plotly.
            figura.write_image(str(ruta_imagen), width=PNG_GRAFICO_ANCHO, height=PNG_GRAFICO_ALTO, scale=1)

            # Inserta bloque visual en el PDF.
            contenido.append(Spacer(1, 10))
            contenido.append(Paragraph(nombre.replace("_", " ").title(), estilos["Heading3"]))
            contenido.append(Image(str(ruta_imagen), width=PDF_GRAFICO_ANCHO, height=PDF_GRAFICO_ALTO))

        # Informa degradación sin bloquear la generación del PDF.
        except Exception as exc:
            print(f"Aviso: no se pudo generar el gráfico '{nombre}' para el PDF: {exc}")
            continue

    # Compone el PDF final en disco.
    documento.build(contenido)


# Genera ficheros HTML/PDF/Excel para el informe premium de GA4.
def generar_informe_ga4_premium(
    configuracion: Configuracion,
    carpeta_salida: Path,
    cliente_nombre: str,
    gestor: str,
    fecha_desde: str,
    fecha_hasta: str,
    comparacion: str,
    provincia: str,
) -> dict[str, Any]:
    """Construye un informe GA4 premium con exportación múltiple y degradación elegante."""

    # Inicializa respuesta por defecto cuando no se pueda consultar GA4.
    salida: dict[str, Any] = {"activo": False, "error": ""}

    # Valida configuración mínima antes de conectar con GA4.
    if not configuracion.ga_enabled or not configuracion.ga_property_id or not configuracion.ga_credentials_file:
        # Registra causa funcional sin lanzar excepción.
        salida["error"] = "GA4 no está disponible en la configuración actual."
        return salida

    # Garantiza existencia de carpeta destino para exportaciones.
    carpeta_salida.mkdir(parents=True, exist_ok=True)

    # Resuelve periodo comparativo según preferencia CLI.
    fecha_comp_desde, fecha_comp_hasta = _resolver_comparacion(fecha_desde, fecha_hasta, comparacion)

    # Inicializa cliente y datasets en bloque controlado.
    try:
        # Crea cliente autenticado para consultas GA4.
        cliente = _crear_cliente(configuracion.ga_credentials_file)

        # Carga datasets necesarios para el informe premium.
        datasets = _cargar_datasets_ga4(
            cliente, configuracion.ga_property_id, fecha_desde, fecha_hasta, fecha_comp_desde, fecha_comp_hasta
        )

    # Maneja cualquier error remoto/auth de forma no bloqueante.
    except Exception as exc:
        # Propaga mensaje de error controlado para CLI.
        salida["error"] = f"No fue posible consultar GA4: {exc}"
        return salida

    # Calcula KPIs y variaciones requeridos por el informe.
    kpis, variaciones = _calcular_kpis(datasets.get("kpi_actual", []), datasets.get("kpi_comparado", []))

    # Construye insights automáticos requeridos.
    insights = _construir_insights(datasets.get("landings", []))

    # Genera gráficos y dataframes base para exportaciones.
    graficos, dataframes = _generar_graficos(datasets)

    # Define rutas de salida por formato.
    ruta_html = carpeta_salida / "informe_ga4_premium.html"
    ruta_excel = carpeta_salida / "informe_ga4_premium.xlsx"
    ruta_pdf = carpeta_salida / "informe_ga4_premium.pdf"

    # Exporta versión HTML interactiva.
    _exportar_html(
        ruta_html,
        cliente_nombre,
        gestor,
        fecha_desde,
        fecha_hasta,
        fecha_comp_desde,
        fecha_comp_hasta,
        kpis,
        variaciones,
        insights,
        graficos,
        dataframes,
        datasets.get("social", []),
        provincia,
    )

    # Exporta versión Excel con dashboard y detalle GA4.
    _exportar_excel(ruta_excel, cliente_nombre, fecha_desde, fecha_hasta, kpis, variaciones, datasets)

    # Exporta versión PDF estática con degradación por gráfico.
    _exportar_pdf(
        ruta_pdf,
        carpeta_salida,
        cliente_nombre,
        gestor,
        fecha_desde,
        fecha_hasta,
        kpis,
        variaciones,
        insights,
        graficos,
    )

    # Rellena metadatos de salida exitosa.
    salida["activo"] = True
    salida["error"] = ""
    salida["html"] = str(ruta_html)
    salida["pdf"] = str(ruta_pdf)
    salida["excel"] = str(ruta_excel)
    salida["insights"] = insights
    return salida
